#!/usr/bin/env python3
"""
Export des données ISS depuis SQLite vers Excel.

Usage:
    python3 scripts/export_excel.py                          # DB par défaut (backend/data/iss.db)
    python3 scripts/export_excel.py --db /path/to/iss.db     # DB custom
    python3 scripts/export_excel.py --output rapport.xlsx    # Fichier de sortie custom
"""

import argparse
import os
import sqlite3
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("openpyxl requis. Installez-le avec : pip3 install openpyxl")
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Export ISS data to Excel")
    parser.add_argument("--db", default=None, help="Path to SQLite database")
    parser.add_argument("--output", "-o", default=None, help="Output Excel file path")
    args = parser.parse_args()

    # Find DB
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    db_path = args.db or os.path.join(project_root, "backend", "data", "iss.db")

    if not os.path.exists(db_path):
        print(f"Base de données introuvable : {db_path}")
        sys.exit(1)

    # Output path
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = args.output or os.path.join(project_root, f"extraction_iss_{timestamp}.xlsx")

    print(f"DB      : {db_path}")
    print(f"Output  : {output_path}")

    db = sqlite3.connect(db_path)
    db.row_factory = sqlite3.Row

    # Last sync info
    sync = db.execute("SELECT * FROM sync_run WHERE status='success' ORDER BY id DESC LIMIT 1").fetchone()
    if sync:
        print(f"Synchro : {sync['finished_at']} — {sync['events_pulled']} events")
    else:
        print("Aucune synchro réussie trouvée.")

    # Data elements
    des = db.execute(
        "SELECT de_uid, code, name, section_prefix FROM metadata_de ORDER BY section_prefix, name"
    ).fetchall()
    de_map = {r["de_uid"]: r["name"] for r in des}
    de_order = [r["de_uid"] for r in des]

    # Events
    events = db.execute(
        "SELECT event_uid, org_unit_uid, org_unit_name, district, region, event_date, status "
        "FROM event ORDER BY region, district, org_unit_name"
    ).fetchall()

    # Values (bulk load)
    vals = db.execute("SELECT event_uid, de_uid, value FROM event_value").fetchall()
    val_map = {}
    for v in vals:
        val_map.setdefault(v["event_uid"], {})[v["de_uid"]] = v["value"]

    # Quality scores
    scores = db.execute(
        "SELECT event_uid, score, n_error, n_warning, n_info FROM event_quality"
    ).fetchall()
    score_map = {r["event_uid"]: dict(r) for r in scores}

    print(f"Events  : {len(events)}")
    print(f"DEs     : {len(de_order)}")
    print(f"Values  : {len(vals)}")
    print("Génération Excel...")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Données ISS"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    fixed_cols = [
        "Région", "District", "Structure", "Org Unit UID", "Event UID",
        "Date", "Statut", "Score Qualité", "Erreurs", "Avertissements", "Infos",
    ]
    de_names = [de_map.get(uid, uid) for uid in de_order]
    all_headers = fixed_cols + de_names

    # Headers
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_idx, evt in enumerate(events, 2):
        eid = evt["event_uid"]
        sc = score_map.get(eid, {})

        ws.cell(row=row_idx, column=1, value=evt["region"])
        ws.cell(row=row_idx, column=2, value=evt["district"])
        ws.cell(row=row_idx, column=3, value=evt["org_unit_name"])
        ws.cell(row=row_idx, column=4, value=evt["org_unit_uid"])
        ws.cell(row=row_idx, column=5, value=eid)
        ws.cell(row=row_idx, column=6, value=evt["event_date"])
        ws.cell(row=row_idx, column=7, value=evt["status"])
        ws.cell(row=row_idx, column=8, value=sc.get("score", ""))
        ws.cell(row=row_idx, column=9, value=sc.get("n_error", ""))
        ws.cell(row=row_idx, column=10, value=sc.get("n_warning", ""))
        ws.cell(row=row_idx, column=11, value=sc.get("n_info", ""))

        event_vals = val_map.get(eid, {})
        for de_idx, de_uid in enumerate(de_order):
            val = event_vals.get(de_uid, "")
            if val:
                col = len(fixed_cols) + de_idx + 1
                try:
                    ws.cell(row=row_idx, column=col, value=float(val))
                except ValueError:
                    ws.cell(row=row_idx, column=col, value=val)

    # Formatting
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{len(events) + 1}"
    ws.freeze_panes = "L2"
    for i, w in enumerate([15, 18, 30, 15, 15, 12, 12, 10, 8, 12, 6], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    db.close()

    size_mb = os.path.getsize(output_path) / (1024 * 1024)
    print(f"OK : {len(events)} lignes, {len(all_headers)} colonnes, {size_mb:.1f} MB")
    print(f"Fichier : {output_path}")


if __name__ == "__main__":
    main()
